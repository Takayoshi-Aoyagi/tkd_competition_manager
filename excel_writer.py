#import openpyxl
from openpyxl import Workbook

from dataclasses import dataclass


@dataclass
class PlayerExcelWriter:
    participants: list
    massogi_map: map
    tul_map: map
    dojo_map: map

    def create_participants_sheet(self, wb, title, participants):
        sheet = wb.create_sheet(title=title)
        row = 1

        headers = ['氏名', 'かな', 'Name', '性別', '級位', '道場名', 'トゥル', 'マッソギ']
        for col, header in enumerate(headers):
            sheet.cell(column=col+1, row=row, value=header)

        for p in participants:
            row += 1
            sheet.cell(column=1, row=row, value=p.name)
            sheet.cell(column=2, row=row, value=p.kana_name)
            sheet.cell(column=3, row=row, value=p.roma_name)
            sheet.cell(column=4, row=row, value=p.gender)
            sheet.cell(column=5, row=row, value=p.degree)
            sheet.cell(column=6, row=row, value=p.dojo)
            sheet.cell(column=7, row=row, value=p.tul)
            sheet.cell(column=8, row=row, value=p.massogi)            
        
    def create_dojo_book(self):
        wb = Workbook()

        participants = []
        for _, ps in self.dojo_map.items():
            participants.extend(ps)
        self.create_participants_sheet(wb, '選手一覧', participants)

        maps = [self.dojo_map]
        for _map in maps:
            for classification, category_participants in _map.items():
                self.create_participants_sheet(wb, classification, category_participants)
        del wb['Sheet'] # remove default sheet
        wb.save(filename='道場別選手一覧.xlsx')
        
    def create_classification_book(self):
        wb = Workbook()

        maps = [self.massogi_map, self.tul_map]
        for _map in maps:
            for classification, category_participants in _map.items():
                self.create_participants_sheet(wb, classification, category_participants)

        del wb['Sheet'] # remove default sheet
        wb.save(filename='競技種目エントリー一覧.xlsx')

    def execute(self):
        self.create_dojo_book()
        self.create_classification_book()        
