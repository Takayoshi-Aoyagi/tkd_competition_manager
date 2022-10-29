from dataclasses import dataclass, field
from functools import cmp_to_key
import glob
import sys

from openpyxl import Workbook, load_workbook
import pandas as pd

from participant import Participant


@dataclass
class ParticipantExcelReader:
    vacant_to_withdraw: bool = False
    errors: list = field(default_factory=list)

    def get_degree(self, val):
        if val == '段、級':
            return None
        kws = ['級', '段']
        for kw in kws:
            if kw in val:
                return val
        return None

    def get_name(self, name):
        if name in ['川口　太郎', 'わらび　花子']:
            return None
        return name

    def get_cell_value(self, cell, dojo, name, column):
        if type(cell) is not str:
            if self.vacant_to_withdraw and pd.isnull(cell):
                return '×'
            err = f'ERROR: Invalid value ({dojo}, {name}, {column}={cell})'
            self.errors.append(err)
        return cell

    def read_row(self, row):
        name = self.get_name(row[1])
        if pd.isnull(name):
            return None
        gender = row[3]
        degree = self.get_degree(row[4])
        if degree is None:
            return None
        dojo = row[5]
        tul = self.get_cell_value(row[6], dojo, name, 'トゥル')
        massogi = self.get_cell_value(row[9], dojo, name, 'マッソギ')
        roma_name = self.get_cell_value(row[13], dojo, name, '英語表記')
        kana_name = row[15]
        # print(row)
        participant = Participant(
            name=name,
            gender=gender,
            degree=degree,
            dojo=dojo,
            tul=tul,
            massogi=massogi,
            roma_name=roma_name,
            kana_name=kana_name
        )
        return participant

    def read_file(self, fpath):
        df = pd.read_excel(fpath)
        participants = []
        for index, row in df.iterrows():
            participant = self.read_row(row)
            if participant is None:
                continue
            participants.append(participant)
        return participants

    def read_files(self, input_dir):
        pattern = f'{input_dir}/*.xlsx'
        all_participants = []
        for fpath in glob.glob(pattern):
            print(f'Reading file: {fpath}')
            participants = self.read_file(fpath)
            all_participants.extend(participants)
            print('Done')
        return all_participants

    def exit_with_error(self):
        print('=============================')
        for err in self.errors:
            print(err)
        sys.exit(1)

    def execute(self, input_dir='input'):
        all_participants = self.read_files(input_dir)
        # print(all_participants)
        if len(self.errors) > 0:
            self.exit_with_error()
        return all_participants


@dataclass
class Result:
    event: str
    classification: str
    rank: str
    name: str


@dataclass
class ResultsExcelReader:
    fpath: str

    def get_results(self):
        COL_1ST_PRIZE = 4
        results = []
        df = pd.read_excel(self.fpath)
        for _, row in df.iterrows():
            # print(row)
            for i in range(COL_1ST_PRIZE, COL_1ST_PRIZE + 4):
                name = row[i]
                if pd.isnull(name):
                    continue
                rank = df.columns[i]
                result = Result(
                    event=row[1],
                    classification=row[3],
                    rank=rank,
                    name=name)
                results.append(result)
                print(result)
        return results

    def execute(self):
        results = self.get_results()
        return results


@dataclass
class ExcelWriter:

    filename: str

    def create_participants_sheet(self, wb, title, participants):
        sheet = wb.create_sheet(title=title)
        row = 1

        headers = ['氏名', 'かな', 'Name', '性別', '級位', '道場名', 'トゥル', 'マッソギ']
        for col, header in enumerate(headers):
            sheet.cell(column=col+1, row=row, value=header)

        for p in participants:
            row += 1
            sheet.cell(column=1, row=row, value=p.name)
            sheet.cell(column=2, row=row, value=p.kana_name)
            sheet.cell(column=3, row=row, value=p.roma_name)
            sheet.cell(column=4, row=row, value=p.gender)
            sheet.cell(column=5, row=row, value=p.degree)
            sheet.cell(column=6, row=row, value=p.dojo)
            sheet.cell(column=7, row=row, value=p.tul)
            sheet.cell(column=8, row=row, value=p.massogi)

    def execute(self):
        wb = Workbook()
        self.create_sheets(wb)
        del wb['Sheet']  # remove default sheet
        wb.save(filename=self.filename)


@dataclass
class DojoExcelWriter(ExcelWriter):
    participants: list
    dojo_map: map

    def create_sheets(self, wb):
        participants = []
        for _, ps in self.dojo_map.items():
            participants.extend(ps)
        self.create_participants_sheet(wb, '選手一覧', participants)

        for classification, category_participants in self.dojo_map.items():
            self.create_participants_sheet(wb, classification,
                                           category_participants)

def sort_by_participahts_desc(e1, e2):
    p_list1 = e1[1]
    p_list2 = e1[1]
    return len(p_list1) - len(p_list2)


@dataclass
class EventExcelWriter(ExcelWriter):
    event_map: map

    def create_sheets(self, wb):
        # sort by number of participants (desc)
        sorted_items = sorted(
            self.event_map.items(), key=cmp_to_key(sort_by_participahts_desc))
        for classification, category_participants in sorted_items:
            self.create_participants_sheet(wb, classification,
                                           category_participants)

def get_tournament_order(num):
    num_order = {
        2: [0, 1],
        4: [0, 2, 1, 3],
        8: [0, 4, 2, 6, 1, 5, 3, 7],
        16: [0, 8, 4, 12, 2, 14, 6, 10, 3, 11, 7, 15, 5, 13, 9, 1],
        32: [0, 16, 8, 24, 12, 28, 4, 20, 14, 30,
             6, 22, 10, 18, 26, 2, 5, 25, 13, 29,
             11, 23, 19, 3, 9, 27, 7, 21, 31, 17,
             15, 1]
    }
    return num_order[num]


@dataclass
class TournamentChartWriter:
    filename: str
    event_map: map
    template_file: str = 'templates/tmpl_tournament.xlsx'
    tmpl_sheet_names: list[int] = field(default_factory=lambda: [2, 4, 8, 16, 32])

    def delete_sheets(self, wb):
        for name in list(map(str, self.tmpl_sheet_names)):
            del wb[name]  # remove default sheet

    def get_sheet(self, wb, num_participants, classification):
        for limit in self.tmpl_sheet_names:
            if num_participants <= limit:
                sheet = wb.copy_worksheet(wb[str(limit)])
                sheet.title = classification
                return sheet, limit
        raise Exception(f'Limit exceeded: {num_participants}: {classification}')

    def write_sheet(self, participants, sheet, sheet_name):
        col_index = 7
        order_list = get_tournament_order(sheet_name)
        for i, p in enumerate(participants):
            order = order_list[i]
            row_index = order * 4 + 2
            name_dojo = f'{p.name} ({p.dojo})'
            sheet.cell(column=col_index, row=row_index, value=name_dojo)
            sheet.cell(column=col_index, row=row_index + 1, value=p.kana_name)

    def create_sheets(self, wb):
        classifications = sorted(self.event_map.keys())
        for classification in classifications:
            if classification == '×':
                continue
            category_participants = self.event_map[classification]
            num_participants = len(category_participants)
            sheet, sheet_name = self.get_sheet(wb, num_participants,
                                               classification)
            self.write_sheet(category_participants, sheet, sheet_name)

    def execute(self):
        wb = load_workbook(self.template_file)
        self.create_sheets(wb)
        self.delete_sheets(wb)
        wb.save(filename=self.filename)
